"""
nsw_licence_scraper.py — Load NSW contractor licence register and enrich builder records.

Data source: NSW Fair Trading / OneGov contractor licence Excel file.
Download URL: http://onegov.nsw.gov.au/agencies/oft/Contractor%20Licence.xlsx
Last file date: December 2023 (updated periodically by NSW Government).

The Excel has 5 sheets:
  Sheet3: Organisation licensees — companies with licence number, address, classes
  Sheet2: Individual licensees — sole traders (optional, filtered by --include-individuals)

This script:
  1. Downloads the Excel (or uses a cached copy)
  2. Filters to active Builder-class licences in NSW
  3. For each: calls enrich_new_candidate() which finds their website via AI-validated
     DuckDuckGo search, then extracts email/owner/type from their website
  4. Only inserts records that pass the 3-field gate: name + website + full street address

Usage:
  python nsw_licence_scraper.py                    # orgs only, all NSW builders
  python nsw_licence_scraper.py --limit 200        # stop after 200 attempts
  python nsw_licence_scraper.py --dry-run          # print records without inserting
  python nsw_licence_scraper.py --refresh          # re-download the Excel file
"""

import argparse
import logging
import os
import re
import sys
import time
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from shared.supabase_client import _get_client

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger(__name__)

EXCEL_URL = 'http://onegov.nsw.gov.au/agencies/oft/Contractor%20Licence.xlsx'
EXCEL_CACHE = os.path.join(os.path.dirname(__file__), '..', 'data', 'nsw_contractor_licence.xlsx')

BUILDER_CLASSES = {'builder', 'general building work', 'full building work', 'building work'}

FRANCHISE_NAMES = {
    'gj gardner', 'masterton', 'metricon', 'henley', 'porter davis',
    'simonds', 'burbank', 'mcdonald jones', 'wisdom homes', 'clarendon',
    'hotondo', 'stroud homes', 'orbit homes', 'plantation homes', 'dale alcock',
}


def _download_excel(force: bool = False) -> str:
    """Download the Excel file if not cached. Returns local path."""
    os.makedirs(os.path.dirname(EXCEL_CACHE), exist_ok=True)

    if os.path.exists(EXCEL_CACHE) and not force:
        log.info(f'Using cached Excel: {EXCEL_CACHE}')
        return EXCEL_CACHE

    log.info(f'Downloading NSW contractor licence Excel from {EXCEL_URL}...')
    import requests
    r = requests.get(EXCEL_URL, timeout=120, stream=True)
    r.raise_for_status()
    with open(EXCEL_CACHE, 'wb') as f:
        for chunk in r.iter_content(chunk_size=65536):
            f.write(chunk)
    size_mb = os.path.getsize(EXCEL_CACHE) / (1024 * 1024)
    log.info(f'Downloaded {size_mb:.1f} MB → {EXCEL_CACHE}')
    return EXCEL_CACHE


def _normalise_address(raw: str) -> str:
    """Convert ALL-CAPS comma-separated address to Title Case for readability."""
    if not raw:
        return ''
    # "89 REX RD,GEORGES HALL,NSW 2198" → "89 Rex Rd, Georges Hall NSW 2198"
    parts = [p.strip().title() for p in raw.split(',')]
    # Keep state abbreviation uppercase: "Nsw" → "NSW"
    result = ', '.join(parts)
    result = re.sub(r'\b(Nsw|Act|Vic|Qld|Sa|Wa|Tas|Nt)\b', lambda m: m.group().upper(), result)
    return result


def _is_franchise(name: str) -> bool:
    n = name.lower()
    return any(f in n for f in FRANCHISE_NAMES)


def load_builders(excel_path: str, include_individuals: bool = False) -> list[dict]:
    """Read the Excel and return list of active NSW builder organisation records."""
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl not installed. Run: pip install openpyxl')
        sys.exit(1)

    today = date.today()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheets_to_read = ['Sheet3']  # Organisation licensees
    if include_individuals:
        sheets_to_read.append('Sheet2')

    results = []
    seen_names: set[str] = set()

    for sheet_name in sheets_to_read:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        is_org = (sheet_name == 'Sheet3')

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue

            licence_num = str(row[0]).strip()
            expiry = row[2]
            name = str(row[3]).strip().strip() if row[3] else ''
            address_raw = str(row[5]).strip() if row[5] else ''
            acn = str(row[7]).strip() if len(row) > 7 and row[7] else None
            abn = str(row[8]).strip() if len(row) > 8 and row[8] else None
            classes = str(row[9]).strip() if len(row) > 9 and row[9] else ''

            # Filter: must be a builder class
            class_lower = classes.lower()
            if not any(bc in class_lower for bc in BUILDER_CLASSES):
                continue

            # Filter: must be NSW address
            if 'NSW' not in address_raw.upper():
                continue

            # Filter: must be active (not expired)
            if isinstance(expiry, datetime):
                expiry_date = expiry.date()
            elif isinstance(expiry, date):
                expiry_date = expiry
            else:
                expiry_date = today  # unknown expiry — assume active
            if expiry_date < today:
                continue

            # Clean name
            name = re.sub(r'\s+', ' ', name).strip()
            if not name or len(name) < 3:
                continue

            # Skip franchises (large national builders — wrong target)
            if _is_franchise(name):
                continue

            name_lower = name.lower()
            if name_lower in seen_names:
                continue
            seen_names.add(name_lower)

            # Normalise address
            address = _normalise_address(address_raw)

            results.append({
                'company_name': name,
                'postal_address': address,
                'builder_licence': licence_num,
                'abn': abn,
                'source': 'nsw_licence',
                'is_org': is_org,
            })

    log.info(f'Loaded {len(results)} active NSW builder records from Excel')
    return results


def run(limit: int = 0, dry_run: bool = False, refresh_excel: bool = False,
        include_individuals: bool = False) -> dict:
    excel_path = _download_excel(force=refresh_excel)
    builders = load_builders(excel_path, include_individuals)

    if not builders:
        log.error('No builders loaded from Excel')
        return {'inserted': 0, 'skipped': 0}

    if dry_run:
        log.info(f'\n[DRY RUN] First 20 of {len(builders)} records:')
        for b in builders[:20]:
            log.info(f"  {b['company_name']} | {b['postal_address']} | {b['builder_licence']}")
        return {'inserted': 0, 'skipped': 0, 'found': len(builders)}

    supabase = _get_client()
    log.info('Loading existing prospect names...')
    resp = supabase.table('builder_prospects').select('company_name').limit(10000).execute()
    existing_lower = {r['company_name'].lower() for r in (resp.data or [])}
    log.info(f'  {len(existing_lower)} existing prospects loaded')

    from acquisition.enricher import enrich_new_candidate

    total_inserted = 0
    total_skipped = 0
    attempts = 0

    for b in builders:
        if limit and attempts >= limit:
            break

        name = b['company_name']
        if name.lower() in existing_lower:
            total_skipped += 1
            continue

        attempts += 1
        log.info(f'[{attempts}] {name} | {b["postal_address"]}')

        candidate = {
            'company_name': name,
            'postal_address': b['postal_address'],  # full street address from register
            'builder_licence': b.get('builder_licence'),
            'abn': b.get('abn'),
            'website': None,  # enricher finds via AI-validated DuckDuckGo
            'source': 'nsw_licence',
        }

        inserted = enrich_new_candidate(supabase, candidate, existing_lower)
        if inserted:
            existing_lower.add(name.lower())
            total_inserted += 1
            if total_inserted % 10 == 0:
                log.info(f'  ✓ {total_inserted} inserted so far')
        else:
            total_skipped += 1

        time.sleep(1.5)

    log.info(f'\nDone — inserted: {total_inserted}, skipped: {total_skipped}')
    return {'inserted': total_inserted, 'skipped': total_skipped}


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Load NSW builder licence register into prospects DB')
    parser.add_argument('--limit', type=int, default=0, help='Max records to attempt (0 = all)')
    parser.add_argument('--dry-run', action='store_true', help='Print records without inserting')
    parser.add_argument('--refresh', action='store_true', help='Re-download the Excel file')
    parser.add_argument('--individuals', action='store_true',
                        help='Also include individual (sole trader) licences')
    args = parser.parse_args()
    run(args.limit, args.dry_run, args.refresh, args.individuals)
